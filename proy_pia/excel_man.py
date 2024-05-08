from openpyxl import load_workbook

book = load_workbook('menus.xlsx')

alph = 'ABCDEFGHIJKLMNOPQRSTUVWXY'

desayuno = []
colacion1 = []
comida = []
colacion2 = []
cena = []

menu1 = []
menu2 = []
menu3 = []
menu4 = []
menu5 = []
menu6 = []

#Por cada DIA hay 5 COMIDAS y cada comida tiene su RECETA
def count():
  for sheets in book:
    for col in range(sheets.min_column, sheets.max_column):
      for row in range(sheets.min_row+1,sheets.max_row+1):
        if sheets[f'{alph[col]}{row}'].value != None:
          text = sheets[f'{alph[col]}{row}'].value
          if alph[col] == 'C':
            desayuno.append(text)
          elif alph[col] == 'D':
            colacion1.append(text)
          elif alph[col] == 'E':
            comida.append(text)
          elif alph[col] == 'F':
            colacion2.append(text)
          elif alph[col] == 'G':
            cena.append(text)

  for sheets in book:
    print(sheets.min_row, sheets.max_row)
  #return menus
    
def chop():
  positions = []
  sort=[]
  for i in range(0,len(desayuno),1):
    if desayuno[i] == "DESAYUNO":
      positions.append(i)

  for j in range(0,len(positions)):
    try:
      sort.append(desayuno[positions[j]:positions[j+1]])
    except IndexError:
      sort.append(desayuno[positions[j]:])
    menu1.append(sort)
  print(menu1[0][0])
  #print(positions)

#Posiciones de DESAYUNO: 0,89,135,170,215

'''
for sheets in book:
  print(sheets.max_row)
  for i in range(sheets.min_row+1, sheets.max_row+1):
    if sheets[f'B{i}'].value != None:
      #print(sheets[f'B{i}'].value)
      days.append(f'B{i}')
      numbers.append(i)

  for j in range(0,sheets.max_column+1):
    letter = sheets[f'{alph[j]}2'].value
    if letter != None and f'{alph[j]}2' not in foods:
      foods.append(f'{alph[j]}2')
  #for k in foods:
    #print(k)
'''

  #for num in range(0,len(numbers)-1):
   #   if numbers[num+1] > numbers[num]:
    #   for item in foods:
      #    print()
      


  
#Acuerdate de hacer lo que sigue, que es restar el numero siguiente al actual, ej: B15 - B4 = 11, entonces esas 11 columnas son las que leera.

count()
print(len(desayuno))
chop()
#print(desayuno)
#[print(i) for i in menus]
#print(book['Hoja1']['C4':'C13'][1][0].value)
#print(days)
#print(foods)
#print(numbers)
#print(len(completo))
#print(book.sheetnames)
